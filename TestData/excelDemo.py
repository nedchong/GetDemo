import openpyxl
import openpyxl as openpyxl

file_name = "C:\\Users\\ned\\PycharmProjects\\PythonSelFramework\\TestData\\PythonDemo.xlsx"
book = openpyxl.load_workbook(file_name)
sheet = book.active

cell = sheet.cell(row=1, column=2)
print(cell.value)

# sheet.cell(row=2, column=2).value = "Ned"
# book.save(file_name)
# print(sheet.cell(row=2, column=2).value)

# print(sheet.max_row)
# print(sheet.max_row)

# print(sheet['B2'].value)

dict = {}
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase1":
        for j in range(2, sheet.max_column + 1):
            # print(sheet.cell(row=i, column=j).value)
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(dict)

