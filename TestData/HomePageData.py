import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname":"Rahul","lastname":"shetty","gender":"Male"}, {"firstname":"Anshika", "lastname":"shetty", "gender":"Female"}]

    @staticmethod
    def getTestData(test_case_name):
        dict = {}
        file_name = "C:\\Users\\ned\\PycharmProjects\\PythonSelFramework\\TestData\\PythonDemo.xlsx"
        book = openpyxl.load_workbook(file_name)
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"] = "ln1"
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[dict]

